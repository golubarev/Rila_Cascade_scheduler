"""
grafik_parser.py
================
The single, shared definition of *how to read a Grafik .xlsm file*.

Both the full loader (load_grafik.py) and the incremental updater
(update_grafik.py) call ``parse_grafik`` here, so the two can never drift apart:
fix a parsing detail once and both benefit.

This module knows nothing about the database - it only turns a workbook into
plain Python objects (DayBlock / hourly rows).  The database side lives in
db_writer.py.
"""

import datetime
import re

import openpyxl

from . import columns as C

# A period label looks like "00:00 - 01:00".
PERIOD_RE = re.compile(r"^\s*\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2}\s*$")


def num(value):
    """Coerce a raw cell to float, or None if it is not a genuine number.

    Blanks, empty strings and error text (e.g. '#REF!') all become None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def as_date(value):
    """Return a datetime.date if the cell holds a date/datetime, else None."""
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    return None


class DayBlock:
    """Everything parsed for one calendar day of the schedule.

    Attributes
    ----------
    date : datetime.date
    k_coeff, target_rila, target_pastra, target_kamenitza, status :
        the per-day header parameters from the block's top rows.
    hours : list of (seq, hour, ts_eet, row_tuple)
        one entry per hourly period, in order.  ``row_tuple`` is the raw cell
        tuple for that hour, from which db_writer pulls the individual values.
    """

    __slots__ = ("date", "k_coeff", "target_rila", "target_pastra",
                 "target_kamenitza", "status", "hours")

    def __init__(self, date):
        self.date = date
        self.k_coeff = None
        self.target_rila = None
        self.target_pastra = None
        self.target_kamenitza = None
        self.status = None
        self.hours = []


def scan_dates(path):
    """Cheap first pass: return the set of calendar dates present in a file.

    Reads only column A, so it is fast even on the large workbooks.  Used by the
    full loader to work out which file owns which day before loading anything.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Grafik"]
    dates = set()
    for (a_cell,) in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        d = as_date(a_cell)
        if d is not None:
            dates.add(d)
    wb.close()
    return dates


def parse_grafik(path):
    """Parse one Grafik workbook into a list of DayBlock objects.

    The block structure of the sheet is:
      * a header row where column A holds the date  (K coefficient sits here),
      * the next row carries the three target levels,
      * then the hourly period rows,
      * then totals / footer rows (ignored).
    We walk the rows once, latching the current date and reading the header
    parameters, then collecting every hourly period row under that date.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Grafik"]

    blocks = []
    current = None            # the DayBlock we are currently filling
    prev_row = None           # remembered so we can read the status stamp above
    pending_targets = False   # set on a date-row, consumed on the row after it

    for row in ws.iter_rows(min_row=1, max_col=66, values_only=True):
        a_date = as_date(row[C.COL_DATE])

        # (a) date in column A -> start a new day block
        if a_date is not None:
            current = DayBlock(a_date)
            blocks.append(current)
            current.k_coeff = num(row[C.COL_K_COEFF])
            if prev_row is not None and isinstance(prev_row[C.COL_STATUS], str):
                current.status = prev_row[C.COL_STATUS]
            pending_targets = True
            prev_row = row
            continue

        # (b) the row right after a date-row holds the target levels
        if pending_targets and current is not None:
            current.target_rila = num(row[C.COL_TARGET_RILA])
            current.target_pastra = num(row[C.COL_TARGET_PASTRA])
            current.target_kamenitza = num(row[C.COL_TARGET_KAMENITZA])
            pending_targets = False
            # not 'continue': a target row is never also an hourly row anyway

        # (c) an hourly period row
        b = row[C.COL_PERIOD]
        if isinstance(b, str) and PERIOD_RE.match(b) and current is not None:
            start_hour = int(b.split(":")[0].strip())
            seq = len(current.hours)  # 0-based position within this day
            ts = f"{current.date.isoformat()} {start_hour:02d}:00"
            current.hours.append((seq, start_hour, ts, row))

        prev_row = row

    wb.close()

    # Keep only blocks that actually have hourly rows (drops stray header dates).
    return [b for b in blocks if b.hours]
