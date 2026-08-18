"""
load_grafik.py
==============
Reads every ``ГРАФИК РАБОТЕН - ...`` Excel file in a folder and loads their
hourly data into the SQLite database defined in database.py.

The job it does, step by step
------------------------------
1.  Discover the source files and work out, for each one, which *year* it is
    for and whether it is the live "ADMIN" file.
2.  First pass - scan only column A of every file to learn which calendar days
    each file contains.  The yearly files deliberately overlap (each carries
    the previous Nov/Dec and part of the next January), so many days appear in
    two files.
3.  Decide a single "winner" file for every day, so each day is loaded exactly
    once (see ``choose_winner``).  Overlaps are reported, never silently merged.
4.  Second pass - stream each file row by row and, for the days it owns, write
    the hourly records and the per-day header parameters into the database.

Running it
----------
    python -m src.load_grafik  --data-dir /path/to/xlsm/folder  --db data/rila_cascade.db
"""

import argparse
import datetime
import glob
import os
import re
import sqlite3
import sys

import openpyxl

# Import our own modules.  This works when the file is run as ``python -m src.load_grafik``.
from . import columns as C
from . import database as db


# A period label looks like "00:00 - 01:00".  This pattern recognises one.
PERIOD_RE = re.compile(r"^\s*\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2}\s*$")

# A year like 2020..2026 appearing in a file name.
YEAR_RE = re.compile(r"(20\d{2})")


def num(value):
    """
    Coerce a raw cell into a float, or return None if it is not a real number.

    The spreadsheet is full of blanks, empty strings and the occasional error
    text (e.g. '#REF!').  We only want genuine numbers; everything else becomes
    NULL in the database.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    # Strings such as "" or "#REF!" are not numbers -> store nothing.
    return None


def as_date(value):
    """Return a datetime.date if the cell holds a date/datetime, else None."""
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    return None


def discover_files(data_dir):
    """
    Find the Grafik files and tag each with (year, is_admin).

    * A file whose name starts with a year (e.g. "2021__ГРАФИК ... 2021.xlsm")
      is a yearly archive; its year is that number.
    * The live file is named "ГРАФИК РАБОТЕН - ADMIN.xlsm" with no leading year;
      we mark it is_admin=True and leave its year as None.
    """
    paths = sorted(glob.glob(os.path.join(data_dir, "*ГРАФИК*.xlsm")))
    files = []
    for p in paths:
        name = os.path.basename(p)
        # A leading year (first 4 chars) marks a yearly archive.
        m = YEAR_RE.match(name)
        if m:
            year, is_admin = int(m.group(1)), False
        else:
            year, is_admin = None, True
        files.append({"path": p, "name": name, "year": year, "is_admin": is_admin})
    return files


def scan_dates(path):
    """First pass: return the set of calendar dates that appear in a file."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Grafik"]
    dates = set()
    for (a_cell,) in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        d = as_date(a_cell)
        if d is not None:
            dates.add(d)
    wb.close()
    return dates


def choose_winner(candidates):
    """
    Given the list of files that all contain a particular day, pick the one
    that should provide that day's data.

    Rule (highest priority first):
      1. a yearly file whose year equals the day's year (the natural "owner"),
      2. otherwise the live ADMIN file (it carries the freshest actuals),
      3. otherwise the file with the most recent year.

    ``candidates`` is a list of (file_dict, day) - all with the same day.
    """
    day = candidates[0][1]
    files = [f for f, _ in candidates]

    def key(f):
        exact_year = 1 if (f["year"] == day.year) else 0
        admin = 1 if f["is_admin"] else 0
        year_rank = f["year"] if f["year"] is not None else -1
        return (exact_year, admin, year_rank)

    return max(files, key=key)


def load(data_dir, db_path):
    """Build the database from scratch out of every Grafik file in ``data_dir``."""
    files = discover_files(data_dir)
    if not files:
        sys.exit(f"No Grafik files found in {data_dir}")

    print(f"Found {len(files)} source file(s):")
    for f in files:
        tag = "ADMIN (live)" if f["is_admin"] else f"year {f['year']}"
        print(f"  - {f['name']}  [{tag}]")

    # --- First pass: which days does each file contain? --------------------
    print("\nScanning dates (first pass)...")
    for f in files:
        f["dates"] = scan_dates(f["path"])
        print(f"  {f['name']}: {len(f['dates'])} days")

    # For every day, gather the files that contain it, then pick one winner.
    day_to_files = {}
    for f in files:
        for d in f["dates"]:
            day_to_files.setdefault(d, []).append((f, d))

    winner_of = {}       # day -> winning file dict
    overlap_days = 0
    for d, cands in day_to_files.items():
        if len(cands) > 1:
            overlap_days += 1
        winner_of[d] = choose_winner(cands)

    print(f"\nTotal distinct days: {len(winner_of)}")
    print(f"Days present in more than one file (overlaps resolved): {overlap_days}")

    # --- Prepare a fresh database -----------------------------------------
    if os.path.exists(db_path):
        os.remove(db_path)                 # rebuild from scratch every run
    os.makedirs(os.path.dirname(db_path) or ".", exist_ok=True)
    conn = db.connect(db_path)
    db.create_schema(conn)

    # Speed pragmas for a big bulk load (safe because we can just re-run).
    conn.execute("PRAGMA synchronous = OFF;")
    conn.execute("PRAGMA journal_mode = MEMORY;")

    # Register each source file and remember its database id.
    now = datetime.datetime.now().isoformat(timespec="seconds")
    for f in files:
        cur = conn.execute(
            "INSERT INTO source_file (filename, file_year, is_admin, imported_at) "
            "VALUES (?,?,?,?)",
            (f["name"], f["year"], 1 if f["is_admin"] else 0, now),
        )
        f["id"] = cur.lastrowid
    conn.commit()

    # --- Second pass: stream each file and load the days it owns -----------
    print("\nLoading rows (second pass)...")
    period_id = 0
    stats = {"periods": 0, "odd_days": []}

    for f in files:
        wb = openpyxl.load_workbook(f["path"], data_only=True, read_only=True)
        ws = wb["Grafik"]

        cur_date = None          # updated whenever column A shows a date
        seq_in_day = 0           # running hour index within the current day
        prev_row = None          # remembered so we can read the status stamp
        pending_targets = False  # set on a date-row, consumed on the next row

        # Batches we flush per file (keeps memory low and inserts fast).
        b_period, b_unit, b_plant, b_res, b_da, b_casc = [], [], [], [], [], []
        b_daily = []
        # Count hours per owned day so we can fill n_hours_in_day at the end.
        hours_per_day = {}

        for row in ws.iter_rows(min_row=1, max_col=66, values_only=True):
            a_date = as_date(row[C.COL_DATE])

            # (a) A date in column A marks the header row of a new daily block.
            if a_date is not None:
                cur_date = a_date
                seq_in_day = 0
                # Capture per-day parameters, but only for days THIS file owns.
                if winner_of.get(cur_date) is f:
                    k_coeff = num(row[C.COL_K_COEFF])
                    status = None
                    if prev_row is not None and isinstance(prev_row[C.COL_STATUS], str):
                        status = prev_row[C.COL_STATUS]
                    # targets live on the *next* row; remember to grab them.
                    b_daily.append([cur_date.isoformat(), k_coeff, None, None, None,
                                    status, f["id"]])
                    pending_targets = True
                prev_row = row
                continue

            # (b) The row immediately after a date-row carries the target levels.
            if pending_targets:
                b_daily[-1][2] = num(row[C.COL_TARGET_RILA])
                b_daily[-1][3] = num(row[C.COL_TARGET_PASTRA])
                b_daily[-1][4] = num(row[C.COL_TARGET_KAMENITZA])
                pending_targets = False
                # (do not 'continue' - a target row is never also an hourly row)

            # (c) An hourly period row - but only load it for owned days.
            b = row[C.COL_PERIOD]
            if isinstance(b, str) and PERIOD_RE.match(b) and cur_date is not None:
                if winner_of.get(cur_date) is not f:
                    prev_row = row
                    continue  # some other file owns this day

                start_hour = int(b.split(":")[0].strip())
                period_id += 1
                pid = period_id
                ts = f"{cur_date.isoformat()} {start_hour:02d}:00"

                b_period.append((pid, ts, cur_date.isoformat(), start_hour,
                                 seq_in_day, None, f["id"]))
                hours_per_day[cur_date] = hours_per_day.get(cur_date, 0) + 1
                seq_in_day += 1
                stats["periods"] += 1

                # unit-level gross generation
                for plant, unit_no, idx in C.UNIT_GROSS:
                    b_unit.append((pid, plant, unit_no, num(row[idx])))

                # plant-level gross + net
                for plant, gidx, nidx in C.PLANT_OUTPUT:
                    b_plant.append((pid, plant, num(row[gidx]), num(row[nidx])))

                # reservoir water-balance state
                for res, idxs in C.RESERVOIR_COLS.items():
                    vals = [num(row[i]) for i in idxs]
                    b_res.append((pid, res, *vals))

                # archived day-ahead schedule per unit
                for plant, unit_no, idx in C.DAYAHEAD_UNIT:
                    b_da.append((pid, plant, unit_no, num(row[idx])))

                # cascade totals / schedule / imbalance / intraday
                b_casc.append((
                    pid,
                    num(row[C.CASCADE["actual_gross"]]), num(row[C.CASCADE["actual_net"]]),
                    num(row[C.CASCADE["submitted_gross"]]), num(row[C.CASCADE["submitted_net"]]),
                    num(row[C.CASCADE["dayahead_gross"]]), num(row[C.CASCADE["dayahead_net"]]),
                    num(row[C.CASCADE["imbalance_gross"]]), num(row[C.CASCADE["imbalance_net"]]),
                    num(row[C.CASCADE["intraday_mwh"]]),
                ))

            prev_row = row

        wb.close()

        # --- flush this file's batches into the database -------------------
        conn.executemany(
            "INSERT OR IGNORE INTO period "
            "(id, ts_eet, date, hour, seq, n_hours_in_day, source_file_id) "
            "VALUES (?,?,?,?,?,?,?)", b_period)
        conn.executemany(
            "INSERT OR IGNORE INTO unit_generation VALUES (?,?,?,?)", b_unit)
        conn.executemany(
            "INSERT OR IGNORE INTO plant_output VALUES (?,?,?,?)", b_plant)
        conn.executemany(
            "INSERT OR IGNORE INTO reservoir_state VALUES (?,?,?,?,?,?,?,?,?)", b_res)
        conn.executemany(
            "INSERT OR IGNORE INTO dayahead_unit VALUES (?,?,?,?)", b_da)
        conn.executemany(
            "INSERT OR IGNORE INTO cascade_balance VALUES (?,?,?,?,?,?,?,?,?,?)", b_casc)
        conn.executemany(
            "INSERT OR IGNORE INTO daily_params "
            "(date, k_source_coeff, target_rila, target_pastra, target_kamenitza, "
            " status, source_file_id) VALUES (?,?,?,?,?,?,?)", b_daily)

        # backfill n_hours_in_day and note any non-24h (DST) days
        for d, n in hours_per_day.items():
            conn.execute("UPDATE period SET n_hours_in_day=? WHERE date=?",
                         (n, d.isoformat()))
            if n != 24:
                stats["odd_days"].append((d.isoformat(), n))

        conn.commit()
        print(f"  loaded from {f['name']}: "
              f"{sum(1 for d in hours_per_day)} owned days")

    conn.close()

    # --- Report ------------------------------------------------------------
    print("\nDone.")
    print(f"  hourly periods loaded: {stats['periods']}")
    print(f"  non-24h (DST) days: {len(stats['odd_days'])} -> "
          f"{sorted(stats['odd_days'])[:6]}{' ...' if len(stats['odd_days'])>6 else ''}")


def main(argv=None):
    ap = argparse.ArgumentParser(description="Load Grafik xlsm files into SQLite.")
    ap.add_argument("--data-dir", required=True, help="folder holding the xlsm files")
    ap.add_argument("--db", required=True, help="path of the SQLite database to build")
    args = ap.parse_args(argv)
    load(args.data_dir, args.db)


if __name__ == "__main__":
    main()